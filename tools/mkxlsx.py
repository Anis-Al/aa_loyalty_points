# -*- coding: utf-8 -*-
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill
from openpyxl.styles import Side
from openpyxl.utils import get_column_letter

P = env['loyalty.program'].browse(2)
rule = P.rule_ids[:1]
rw = P.reward_ids[:1]
cards = env['loyalty.card'].with_context(active_test=False).search([('program_id', '=', 2)])
total_points = round(sum(cards.mapped('points')), 2)
total_money = round(total_points * rw.discount, 2)


def money(v):
    return u'%s %s' % (v, P.currency_id.name)


NAVY = '12305A'
GREY = '6E7C8C'
head_fill = PatternFill('solid', fgColor=NAVY)
head_font = Font(color='FFFFFF', bold=True, size=11)
chg_fill = PatternFill('solid', fgColor='FFF2CC')
sec_fill = PatternFill('solid', fgColor='EEF3FA')
thin = Side(style='thin', color='D7E0EC')
border = Border(left=thin, right=thin, top=thin, bottom=thin)
wrap = Alignment(vertical='top', wrap_text=True)

wb = Workbook()


def sheet(title, widths, rows, intro=None):
    ws = wb.create_sheet(title)
    r = 1
    if intro:
        ws.cell(r, 1, intro).font = Font(italic=True, color=GREY, size=10)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(widths))
        ws.cell(r, 1).alignment = wrap
        ws.row_dimensions[r].height = 26
        r += 2
    head = [u"Field (Odoo label)", u"Technical name", u"Value to set", u"Differs?", u"Note"]
    for c, h in enumerate(head, 1):
        cell = ws.cell(r, c, h)
        cell.fill, cell.font, cell.border = head_fill, head_font, border
    ws.freeze_panes = ws.cell(r + 1, 1)
    r += 1
    for row in rows:
        if len(row) == 1:
            cell = ws.cell(r, 1, row[0])
            cell.font = Font(bold=True, color=NAVY)
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(head))
            for c in range(1, len(head) + 1):
                ws.cell(r, c).fill = sec_fill
                ws.cell(r, c).border = border
        else:
            for c, v in enumerate(row, 1):
                cell = ws.cell(r, c, v)
                cell.border, cell.alignment = border, wrap
                if c == 3:
                    cell.font = Font(bold=True)
                if row[3] == u"YES":
                    cell.fill = chg_fill
        r += 1
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    return ws


ws = wb.active
ws.title = u"Read me"
lines = [
    (u"Loyalty Cards program — setup sheet", 16, True, NAVY),
    (u"", 11, False, GREY),
    (u"Sales › Products › Discount & Loyalty › New — pick “Loyalty Cards”,", 11, False, '000000'),
    (u"then enter the values on the next three tabs.", 11, False, '000000'),
    (u"", 11, False, GREY),
    (u"Every value is copied from the program running today: “%s” (id 2, Next Order Coupons)." % P.name, 11, False, '000000'),
    (u"Only the program type differs. That row is highlighted.", 11, False, '000000'),
    (u"", 11, False, GREY),
    (u"What changes for the customer", 13, True, NAVY),
    (u"Today every order mints a NEW coupon code, so points sit on separate codes and never add up.", 11, False, '000000'),
    (u"A Loyalty Cards program is nominative: one card per customer, reused on every order, so the", 11, False, '000000'),
    (u"balance accumulates. 48 customers currently hold between 2 and 5 separate codes.", 11, False, '000000'),
    (u"", 11, False, GREY),
    (u"Keeping Applies On = Future means points are still earned on one order and spent on a later", 11, False, '000000'),
    (u"one, exactly as today. Setting it to Both would let an order discount itself with the points", 11, False, '000000'),
    (u"it is earning — a business change, not a like-for-like move.", 11, False, '000000'),
    (u"", 11, False, GREY),
    (u"The old coupons", 13, True, NAVY),
    (u"The %s existing codes do not carry over. %s of them hold a balance — %s points across" % (
        len(cards), len(cards.filtered(lambda c: c.points > 0)), total_points), 11, False, '000000'),
    (u"336 customers, worth %s at %s per point." % (money(total_money), money(rw.discount)), 11, False, '000000'),
    (u"", 11, False, GREY),
    (u"Archive the old program once the new one is live, so it stops issuing codes. Archive rather", 11, False, '000000'),
    (u"than delete: deleting a card cascades away its history rows, and 19 of the codes are attached", 11, False, '000000'),
    (u"to confirmed order lines and cannot be deleted at all.", 11, False, '000000'),
]
for i, (text, size, bold, colour) in enumerate(lines, 1):
    ws.cell(i, 1, text).font = Font(size=size, bold=bold, color=colour)
ws.column_dimensions['A'].width = 103

sheet(u"1 - Program", [30, 24, 32, 10, 56], [
    (u"Identification",),
    (u"Program Name", u"name", P.name, u"no", u"Rename freely, nothing keys off it."),
    (u"Program Type", u"program_type", u"Loyalty Cards", u"YES",
     u"The only value that differs. This is what makes the card nominative: one card per customer, reused, instead of one per order."),
    (u"Company", u"company_id", P.company_id.name, u"no", u""),
    (u"Currency", u"currency_id", P.currency_id.name, u"no", u""),
    (u"Behaviour",),
    (u"Applies On", u"applies_on", u"Future orders", u"no",
     u"Keep Future. With Loyalty Cards the program is still nominative, and points stay spendable only on a later order, same as today."),
    (u"Trigger", u"trigger", u"Automatic", u"no", u""),
    (u"Start Date", u"date_from", P.date_from or u"(empty)", u"no", u""),
    (u"End Date", u"date_to", P.date_to or u"(empty)", u"no", u""),
    (u"Limit Usage", u"limit_usage", u"ticked" if P.limit_usage else u"unticked", u"no", u""),
    (u"Pricelist", u"pricelist_ids",
     u", ".join(P.pricelist_ids.mapped('name')) or u"(empty)", u"no", u"Empty means every pricelist."),
    (u"Availability",),
    (u"Sales", u"sale_ok", u"ticked" if P.sale_ok else u"unticked", u"no", u""),
    (u"Point of Sale", u"pos_ok", u"ticked" if P.pos_ok else u"unticked", u"no", u""),
    (u"Available on Website", u"ecommerce_ok", u"ticked" if P.ecommerce_ok else u"unticked", u"no", u""),
    (u"Website", u"website_id", P.website_id.name or u"(all)", u"no", u""),
    (u"Portal",),
    (u"Portal Visible", u"portal_visible", u"ticked" if P.portal_visible else u"unticked", u"no",
     u"Required for /my/loyalty to list the card."),
    (u"Portal Point Name", u"portal_point_name", P.portal_point_name, u"no", u"The unit the customer sees."),
    (u"Communication",),
    (u"Email template", u"mail_template_id", P.mail_template_id.name or u"(none)", u"check",
     u"A new Loyalty Cards program ships with an empty communication plan. Add this template only if the card should be emailed on creation."),
    (u"Credit notes (this module)",),
    (u"Points on Credit Note", u"refund_policy",
     dict(P._fields['refund_policy'].selection)[P.refund_policy], u"no",
     u"Proportional is the default on a new program, which matches today."),
    (u"Allow a Negative Balance", u"allow_negative_points",
     u"ticked" if P.allow_negative_points else u"unticked", u"no", u""),
], intro=u"Program form, top section.")

sheet(u"2 - Conditions", [30, 24, 32, 10, 56], [
    (u"Conditional rule — one line",),
    (u"Application", u"mode", u"Automatic", u"no", u""),
    (u"Minimum Purchase", u"minimum_amount", money(rule.minimum_amount), u"no",
     u"Tax %s." % rule.minimum_amount_tax_mode),
    (u"Minimum Quantity", u"minimum_qty", rule.minimum_qty, u"no", u""),
    (u"Products", u"product_ids",
     u", ".join(rule.product_ids.mapped('name')) or u"(all)", u"no", u"Empty means every product earns."),
    (u"Categories", u"product_category_id", rule.product_category_id.name or u"(all)", u"no", u""),
    (u"Earning rate",),
    (u"Reward Point Mode", u"reward_point_mode", u"per Money spent", u"no",
     u"This is what makes it one point per unit of currency."),
    (u"Reward", u"reward_point_amount", rule.reward_point_amount, u"no",
     u"%s point per %s spent." % (rule.reward_point_amount, P.currency_id.name)),
    (u"Split per unit", u"reward_point_split",
     u"ticked" if rule.reward_point_split else u"unticked", u"no",
     u"Must stay unticked. Ticking it splits the points across several cards, which is the behaviour being moved away from."),
], intro=u"Conditions tab. The live program has exactly one rule.")

sheet(u"3 - Reward", [30, 24, 32, 10, 56], [
    (u"Reward — one line",),
    (u"Reward Type", u"reward_type", u"Discount", u"no", u""),
    (u"Discount Mode", u"discount_mode", u"per Point", u"no",
     u"Per-point is what gives a point a money value. The module prints it as “430 points (4,30 %s)” on the portal, the PDF and the email." % P.currency_id.symbol),
    (u"Discount", u"discount", u"%s %s per point" % (rw.discount, P.currency_id.name), u"no",
     u"The live rate. 100 points = %s." % money(round(100 * rw.discount, 2))),
    (u"Applies to", u"discount_applicability", u"Order", u"no", u""),
    (u"Points needed", u"required_points", rw.required_points, u"no",
     u"1 means the balance is spendable continuously, with no threshold."),
    (u"Max Discount", u"discount_max_amount",
     money(rw.discount_max_amount) if rw.discount_max_amount else u"(none)", u"no",
     u"Leave empty. A cap here would make the money value printed on the card larger than a single order can actually redeem."),
    (u"Description", u"description", rw.description, u"no", u"Printed on the reward order line."),
], intro=u"Rewards tab. The live program has exactly one reward.")

path = 'C:/Program Files/Odoo 19.0.20260724/server/odoo/mnt/aa_loyalty_points/loyalty_program_setup.xlsx'
wb.save(path)
print("saved %s" % path)
