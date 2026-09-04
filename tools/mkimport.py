# -*- coding: utf-8 -*-
"""Build the loyalty.program import file, then prove it through env[...].load()."""
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

P = env['loyalty.program'].browse(2)
rule = P.rule_ids[:1]
rw = P.reward_ids[:1]

COLUMNS = [
    ('id', 'loyalty_program_mconfort'),
    ('name', P.name),
    ('program_type', 'loyalty'),
    ('applies_on', 'future'),
    ('trigger', 'auto'),
    ('company_id', P.company_id.name),
    ('currency_id', P.currency_id.name),
    ('portal_visible', 'True' if P.portal_visible else 'False'),
    ('portal_point_name', P.portal_point_name),
    ('sale_ok', 'True' if P.sale_ok else 'False'),
    ('pos_ok', 'True' if P.pos_ok else 'False'),
    ('ecommerce_ok', 'True' if P.ecommerce_ok else 'False'),
    ('limit_usage', 'True' if P.limit_usage else 'False'),
    ('refund_policy', P.refund_policy),
    ('allow_negative_points', 'True' if P.allow_negative_points else 'False'),
    ('rule_ids/mode', rule.mode),
    ('rule_ids/reward_point_mode', rule.reward_point_mode),
    ('rule_ids/reward_point_amount', rule.reward_point_amount),
    ('rule_ids/reward_point_split', 'True' if rule.reward_point_split else 'False'),
    ('rule_ids/minimum_amount', rule.minimum_amount),
    ('rule_ids/minimum_qty', rule.minimum_qty),
    ('reward_ids/reward_type', rw.reward_type),
    ('reward_ids/discount_mode', rw.discount_mode),
    ('reward_ids/discount', rw.discount),
    ('reward_ids/discount_applicability', rw.discount_applicability),
    ('reward_ids/required_points', rw.required_points),
    ('reward_ids/description', rw.description),
]

fields = [c for c, _ in COLUMNS]
values = [v for _, v in COLUMNS]

wb = Workbook()
ws = wb.active
ws.title = u"loyalty.program"
head_fill = PatternFill('solid', fgColor='12305A')
for c, (name, value) in enumerate(COLUMNS, 1):
    h = ws.cell(1, c, name)
    h.fill = head_fill
    h.font = Font(color='FFFFFF', bold=True, size=10)
    h.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.cell(2, c, value)
    ws.column_dimensions[get_column_letter(c)].width = max(12, min(30, len(name) + 4))
ws.row_dimensions[1].height = 32
ws.freeze_panes = 'A2'

path = 'C:/Program Files/Odoo 19.0.20260724/server/odoo/mnt/aa_loyalty_points/loyalty_program_import.xlsx'
wb.save(path)
print(u"saved %s" % path)

# ---- prove it through the real import path -------------------------------
res = env['loyalty.program'].load(fields, [[str(v) for v in values]])
assert not res['messages'], u"import reported: %s" % res['messages']
new = env['loyalty.program'].browse(res['ids'])
nr, nw = new.rule_ids[:1], new.reward_ids[:1]

checks = [
    ('program_type', new.program_type, 'loyalty'),
    ('applies_on', new.applies_on, 'future'),
    ('is_nominative', new.is_nominative, True),
    ('trigger', new.trigger, P.trigger),
    ('company', new.company_id, P.company_id),
    ('currency', new.currency_id, P.currency_id),
    ('portal_visible', new.portal_visible, P.portal_visible),
    ('portal_point_name', new.portal_point_name, P.portal_point_name),
    ('refund_policy', new.refund_policy, P.refund_policy),
    ('rule count', len(new.rule_ids), 1),
    ('reward count', len(new.reward_ids), 1),
    ('rule.reward_point_mode', nr.reward_point_mode, rule.reward_point_mode),
    ('rule.reward_point_amount', nr.reward_point_amount, rule.reward_point_amount),
    ('rule.minimum_amount', nr.minimum_amount, rule.minimum_amount),
    ('rule.minimum_qty', nr.minimum_qty, rule.minimum_qty),
    ('rule.reward_point_split', nr.reward_point_split, rule.reward_point_split),
    ('reward.reward_type', nw.reward_type, rw.reward_type),
    ('reward.discount_mode', nw.discount_mode, rw.discount_mode),
    ('reward.discount', nw.discount, rw.discount),
    ('reward.discount_applicability', nw.discount_applicability, rw.discount_applicability),
    ('reward.required_points', nw.required_points, rw.required_points),
    ('reward.description', nw.description, rw.description),
]
bad = [(n, g, e) for n, g, e in checks if g != e]
for n, g, e in checks:
    print(u"  %-30s %-22s %s" % (n, g, 'OK' if g == e else u'MISMATCH, expected %s' % (e,)))
print(u"---")
print(u"imported id      : %s" % res['ids'])
print(u"messages         : %s" % res['messages'])
print(u"mismatches       : %s" % len(bad))

# re-import must update, not duplicate
res2 = env['loyalty.program'].load(fields, [[str(v) for v in values]])
print(u"re-import ids    : %s (same record = idempotent)" % res2['ids'])
print(u"programs named   : %s" % env['loyalty.program'].search_count([('name', '=', P.name)]))
assert res2['ids'] == res['ids'], u"re-import created a duplicate"
assert not bad, u"import did not reproduce the live configuration"
env.cr.rollback()
print(u"rolled back - nothing written to the database")
